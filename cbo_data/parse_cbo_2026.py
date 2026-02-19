"""Parse CBO January 2026 Excel files and output updated YAML values.

Usage:
    1. Download CBO Excel files to this directory (cbo_data/)
    2. Run: python cbo_data/parse_cbo_2026.py

Required files (download from https://www.cbo.gov/data/budget-economic-data
and https://www.cbo.gov/data/baseline-projections-selected-programs):
    - Budget projections: 51118-*-Budget-Projections.xlsx
    - Revenue projections: 51138-*-Revenue*.xlsx
    - Economic projections: 51135-*-Economic-Projections.xlsx
    - Tax parameters: 53724-*-Tax-Parameters.xlsx
    - SNAP: 51312-*-snap.xlsx
    - SSI: 51313-*-ssi.xlsx
    - Unemployment: 51316-*-unemployment.xlsx
"""

import glob
import os
import sys

import openpyxl

DATA_DIR = os.path.dirname(os.path.abspath(__file__))


def find_file(pattern):
    """Find a file matching the pattern in the data directory."""
    matches = glob.glob(os.path.join(DATA_DIR, pattern))
    if not matches:
        print(f"WARNING: No file found matching {pattern}")
        return None
    return matches[0]


def read_xlsx(path, sheet_name=None, sheet_index=None):
    """Read an Excel file and return the specified sheet."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name:
        for name in wb.sheetnames:
            if sheet_name.lower() in name.lower():
                return wb[name]
        print(f"  Available sheets: {wb.sheetnames}")
        raise ValueError(f"Sheet '{sheet_name}' not found in {path}")
    if sheet_index is not None:
        return wb[wb.sheetnames[sheet_index]]
    return wb[wb.sheetnames[0]]


def find_row_by_label(sheet, label, col=1, start_row=1, end_row=200):
    """Find the row number containing the given label."""
    for row in range(start_row, end_row + 1):
        cell_val = sheet.cell(row=row, column=col).value
        if cell_val and label.lower() in str(cell_val).lower():
            return row
    return None


def get_row_values(sheet, row, start_col, end_col):
    """Get values from a row across columns."""
    values = []
    for col in range(start_col, end_col + 1):
        val = sheet.cell(row=row, column=col).value
        values.append(val)
    return values


def get_fiscal_year_headers(sheet, header_row, start_col, end_col):
    """Get fiscal year headers from a row."""
    years = []
    for col in range(start_col, end_col + 1):
        val = sheet.cell(row=header_row, column=col).value
        if val:
            try:
                years.append(int(val))
            except (ValueError, TypeError):
                years.append(None)
        else:
            years.append(None)
    return years


def format_dollars(val, unit="dollars"):
    """Format a dollar value for YAML."""
    if val is None:
        return None
    if unit == "billions":
        # Convert billions to dollars
        amount = int(round(val * 1_000_000_000))
    elif unit == "millions":
        amount = int(round(val * 1_000_000))
    else:
        amount = int(round(val))
    # Format with underscores for readability
    s = str(amount)
    # Add underscores every 3 digits from right
    parts = []
    while s:
        parts.append(s[-3:])
        s = s[:-3]
    return "_".join(reversed(parts))


def print_yaml_values(years, values, unit="billions", indent=2):
    """Print YAML-formatted values."""
    prefix = " " * indent
    for year, val in zip(years, values):
        if year and val is not None:
            formatted = format_dollars(val, unit)
            print(f"{prefix}{year}-01-01: {formatted}")


def parse_budget_projections():
    """Parse budget projections for Social Security."""
    path = find_file("51118-*-Budget-Projections.xlsx")
    if not path:
        return
    print(f"\n{'='*60}")
    print(f"BUDGET PROJECTIONS: {os.path.basename(path)}")
    print(f"{'='*60}")

    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    # Look for Social Security in the mandatory spending tables
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        row = find_row_by_label(sheet, "Social Security")
        if row:
            print(
                f"\n  Found 'Social Security' in sheet '{sheet_name}', row {row}"
            )
            # Print surrounding context
            for r in range(max(1, row - 2), row + 3):
                vals = [
                    sheet.cell(row=r, column=c).value for c in range(1, 20)
                ]
                print(f"    Row {r}: {vals}")
            break


def parse_revenue_projections():
    """Parse revenue projections for income tax, payroll taxes, and income by source."""
    path = find_file("51138-*-Revenue*.xlsx")
    if not path:
        return
    print(f"\n{'='*60}")
    print(f"REVENUE PROJECTIONS: {os.path.basename(path)}")
    print(f"{'='*60}")

    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n  --- Sheet: {sheet_name} ---")

        # Look for key rows
        for label in [
            "Individual income tax receipts",
            "individual income",
            "Wages and salaries",
            "Taxable interest",
            "Dividends",
            "Capital gains",
            "Business income",
            "Pension",
            "Social Security",
            "Adjusted gross income",
            "Payroll tax",
            "Social insurance",
        ]:
            row = find_row_by_label(sheet, label)
            if row:
                vals = [
                    sheet.cell(row=row, column=c).value for c in range(1, 20)
                ]
                print(f"    Row {row} ({label}): {vals}")


def parse_economic_projections():
    """Parse economic projections for CPI."""
    path = find_file("51135-*-Economic-Projections.xlsx")
    if not path:
        return
    print(f"\n{'='*60}")
    print(f"ECONOMIC PROJECTIONS: {os.path.basename(path)}")
    print(f"{'='*60}")

    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        # Look for CPI rows
        for label in ["CPI-U", "CPI-W", "Consumer Price Index"]:
            row = find_row_by_label(sheet, label, end_row=300)
            if row:
                vals = [
                    sheet.cell(row=row, column=c).value for c in range(1, 20)
                ]
                print(f"  Sheet '{sheet_name}', Row {row} ({label}): {vals}")


def parse_tax_parameters():
    """Parse tax parameters for C-CPI-U."""
    path = find_file("53724-*-Tax-Parameters.xlsx")
    if not path:
        return
    print(f"\n{'='*60}")
    print(f"TAX PARAMETERS: {os.path.basename(path)}")
    print(f"{'='*60}")

    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for label in ["C-CPI-U", "Chained CPI", "chained"]:
            row = find_row_by_label(sheet, label, end_row=300)
            if row:
                vals = [
                    sheet.cell(row=row, column=c).value for c in range(1, 20)
                ]
                print(f"  Sheet '{sheet_name}', Row {row} ({label}): {vals}")


def parse_snap():
    """Parse SNAP baseline projections."""
    path = find_file("51312-*-snap.xlsx")
    if not path:
        return
    print(f"\n{'='*60}")
    print(f"SNAP: {os.path.basename(path)}")
    print(f"{'='*60}")

    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for label in ["Total Benefits", "Budget Authority", "Total"]:
            row = find_row_by_label(sheet, label)
            if row:
                vals = [
                    sheet.cell(row=row, column=c).value for c in range(1, 20)
                ]
                print(f"  Sheet '{sheet_name}', Row {row} ({label}): {vals}")


def parse_ssi():
    """Parse SSI baseline projections."""
    path = find_file("51313-*-ssi.xlsx")
    if not path:
        return
    print(f"\n{'='*60}")
    print(f"SSI: {os.path.basename(path)}")
    print(f"{'='*60}")

    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for label in ["Estimated Outlays", "Demonstration Projects", "Total"]:
            row = find_row_by_label(sheet, label)
            if row:
                vals = [
                    sheet.cell(row=row, column=c).value for c in range(1, 20)
                ]
                print(f"  Sheet '{sheet_name}', Row {row} ({label}): {vals}")


def parse_unemployment():
    """Parse unemployment compensation baseline projections."""
    path = find_file("51316-*-unemployment.xlsx")
    if not path:
        return
    print(f"\n{'='*60}")
    print(f"UNEMPLOYMENT: {os.path.basename(path)}")
    print(f"{'='*60}")

    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for label in ["Estimated Outlays", "Total"]:
            row = find_row_by_label(sheet, label)
            if row:
                vals = [
                    sheet.cell(row=row, column=c).value for c in range(1, 20)
                ]
                print(f"  Sheet '{sheet_name}', Row {row} ({label}): {vals}")


if __name__ == "__main__":
    print("CBO January 2026 Data Parser")
    print("=" * 60)
    print(f"Looking for files in: {DATA_DIR}\n")

    files = os.listdir(DATA_DIR)
    xlsx_files = [f for f in files if f.endswith(".xlsx")]
    print(f"Found {len(xlsx_files)} Excel files: {xlsx_files}\n")

    if not xlsx_files:
        print(
            "No Excel files found! Please download the CBO data files to this directory."
        )
        print(
            "\nRequired files from https://www.cbo.gov/data/budget-economic-data:"
        )
        print("  - Budget projections (51118-*-Budget-Projections.xlsx)")
        print("  - Revenue projections (51138-*-Revenue*.xlsx)")
        print("  - Economic projections (51135-*-Economic-Projections.xlsx)")
        print("  - Tax parameters (53724-*-Tax-Parameters.xlsx)")
        print(
            "\nFrom https://www.cbo.gov/data/baseline-projections-selected-programs:"
        )
        print("  - SNAP (51312-*-snap.xlsx)")
        print("  - SSI (51313-*-ssi.xlsx)")
        print("  - Unemployment (51316-*-unemployment.xlsx)")
        sys.exit(1)

    parse_budget_projections()
    parse_revenue_projections()
    parse_economic_projections()
    parse_tax_parameters()
    parse_snap()
    parse_ssi()
    parse_unemployment()
